"""Import odpracovaných hodin z Excelů Směny <prodejna>.xlsx pro průměr dovolené."""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, time, timedelta
from pathlib import Path

MONTH_NAMES = {
    'leden': 1,
    'únor': 2,
    'unor': 2,
    'úno': 2,
    'březen': 3,
    'brezen': 3,
    'duben': 4,
    'květen': 5,
    'kveten': 5,
    'červen': 6,
    'cerven': 6,
}

SUMMARY_LABELS = {'celkem', 'součet', 'soucet'}
SKIP_TOKENS = {
    'den', 'jméno', 'jmeno', 'celkem', 'fond', 'rozdíl', 'rozdil', 'svátek', 'svatek',
    'zavřeno', 'zavreno', 'hod.', 'hodiny', 'příchod', 'prichod', 'odchod', 'přestávka',
    'prodejna', 'součet', 'soucet',
}

ALIASES = {
    'benny': 'babušík',
    'králík': 'králik',
    'křížková': 'křížková f',
    'smčková': 'smrčková',
    'smckova': 'smrčková',
}


def _num(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip().lower() in SKIP_TOKENS:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def surname_key(name):
    if name is None:
        return None
    raw = str(name).strip().lstrip('*')
    if not raw:
        return None
    low = raw.lower()
    if any(x in low for x in ('nemoc', 'dovol', 'svátek', 'svatek', 'zavřeno', 'zavreno')):
        return None
    if low in SKIP_TOKENS:
        return None
    head = re.split(r'[-–(]', raw, maxsplit=1)[0].strip()
    parts = [p for p in re.split(r'\s+', head) if p]
    if not parts:
        return None
    token = parts[-1].lower() if len(parts) > 1 else parts[0].lower()
    token = re.sub(r'[^a-záčďéěíňóřšťúůýž\- ]', '', token).strip()
    if not token or token in SKIP_TOKENS or token in MONTH_NAMES:
        return None
    return ALIASES.get(token, token)


def _month_from_cell(value):
    if not isinstance(value, str):
        return None
    return MONTH_NAMES.get(value.strip().lower())


def _shift_header_starts(ws, row):
    den_cols = [
        col for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row, col).value, str)
        and ws.cell(row, col).value.strip().lower() == 'den'
    ]
    if den_cols:
        return den_cols
    jmeno_cols = [
        col for col in range(1, ws.max_column + 1)
        if isinstance(ws.cell(row, col).value, str)
        and ws.cell(row, col).value.strip().lower().startswith('jméno')
    ]
    if jmeno_cols:
        return [col - 1 for col in jmeno_cols if col > 1]
    return []


def _is_shift_header(ws, row):
    return bool(_shift_header_starts(ws, row))


def _hours_col_for_block(ws, header_row, start):
    for off in range(3, 8):
        col = start + off
        if col > ws.max_column:
            break
        val = ws.cell(header_row, col).value
        if isinstance(val, str):
            low = val.strip().lower()
            if low.startswith('hod') or 'hodin' in low:
                return col
    return start + 6


def _names_for_blocks(ws, label_row, starts):
    names = {}
    for start in starts:
        for col in (start, start + 1, start + 2):
            if col < 1 or col > ws.max_column:
                continue
            key = surname_key(ws.cell(label_row, col).value)
            if key:
                names[start] = key
                break
    return names


def _row_hours(ws, row, start, hours_col):
    h = _num(ws.cell(row, hours_col).value)
    if h is not None:
        return max(0.0, h)
    od = ws.cell(row, start + 3).value
    do = ws.cell(row, start + 4).value
    if isinstance(od, (int, float)) and isinstance(do, (int, float)):
        return max(0.0, float(do) - float(od))
    computed = _hours_from_times(od, do) if isinstance(od, time) and isinstance(do, time) else None
    return max(0.0, computed) if computed is not None else None


def _hours_from_times(start, end):
    if not isinstance(start, time) or not isinstance(end, time):
        return None
    s = datetime.combine(datetime.min.date(), start)
    e = datetime.combine(datetime.min.date(), end)
    if e < s:
        e += timedelta(days=1)
    return round((e - s).total_seconds() / 3600, 2)


def _row_level_totals(ws):
    totals = defaultdict(float)
    active = None
    for row in range(1, ws.max_row + 1):
        if _is_shift_header(ws, row):
            starts = _shift_header_starts(ws, row)
            hours_cols = [_hours_col_for_block(ws, row, s) for s in starts]
            active = (starts, hours_cols)
            continue
        if not active:
            continue
        starts, hours_cols = active
        if _is_shift_header(ws, row):
            continue
        for start, hours_col in zip(starts, hours_cols):
            dt = ws.cell(row, start).value
            if not isinstance(dt, datetime):
                continue
            key = surname_key(ws.cell(row, start + 1).value)
            if not key:
                continue
            h = _row_hours(ws, row, start, hours_col)
            if h is None:
                continue
            totals[(dt.year, dt.month, key)] += h
    return totals


def _merge_celkem_and_rows(celkem, rows):
    merged = defaultdict(float)
    celkem_month_users = defaultdict(set)
    for (year, month, surname) in celkem:
        celkem_month_users[(year, month)].add(surname)
    for key, hours in celkem.items():
        merged[key] += hours
    for key, hours in rows.items():
        year, month, surname = key
        if surname in celkem_month_users[(year, month)]:
            continue
        merged[key] += hours
    return merged
    for col in range(1, ws.max_column + 1):
        month = _month_from_cell(ws.cell(row, col).value)
        if month:
            return month
    return None


def _month_on_row(ws, row):
    for col in range(1, ws.max_column + 1):
        month = _month_from_cell(ws.cell(row, col).value)
        if month:
            return month
    return None


def _summary_hours(ws, row, starts, hours_cols):
    out = {}
    for start, hours_col in zip(starts, hours_cols):
        label = ws.cell(row, hours_col - 1).value
        if isinstance(label, str) and label.strip().lower() in SUMMARY_LABELS:
            h = _num(ws.cell(row, hours_col).value)
            if h is not None:
                out[start] = h
    if out:
        return out
    for start, hours_col in zip(starts, hours_cols):
        if start in out:
            continue
        h = _num(ws.cell(row, hours_col).value)
        if h is None:
            continue
        next_label = ws.cell(row + 1, hours_col - 1).value if row + 1 <= ws.max_row else None
        if isinstance(next_label, str) and next_label.strip().lower() == 'fond':
            out[start] = h
    bare = True
    for start in starts:
        if isinstance(ws.cell(row, start).value, datetime):
            bare = False
            break
    if bare:
        numeric = 0
        for hours_col in hours_cols:
            if _num(ws.cell(row, hours_col).value) is not None:
                numeric += 1
        if numeric >= max(1, len(starts) // 2):
            for start, hours_col in zip(starts, hours_cols):
                if start in out:
                    continue
                h = _num(ws.cell(row, hours_col).value)
                if h is not None:
                    out[start] = h
    return out


def _first_date_in_section(ws, row, starts):
    for rr in range(row, min(row + 40, ws.max_row + 1)):
        for start in starts:
            val = ws.cell(rr, start).value
            if isinstance(val, datetime):
                return val.year, val.month
    return None, None


def parse_smeny_workbook(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    celkem = _parse_celkem_totals(ws)
    rows = _row_level_totals(ws)
    return _merge_celkem_and_rows(celkem, rows)


def _parse_celkem_totals(ws):
    totals = defaultdict(float)

    row = 1
    while row <= ws.max_row:
        if not _is_shift_header(ws, row):
            row += 1
            continue

        starts = _shift_header_starts(ws, row)
        if not starts:
            row += 1
            continue

        hours_cols = [_hours_col_for_block(ws, row, start) for start in starts]
        label_row = max(1, row - 1)
        names = _names_for_blocks(ws, label_row, starts)
        if not names and row - 2 >= 1:
            names = _names_for_blocks(ws, row - 2, starts)

        month = _month_on_row(ws, label_row)
        if month is None and row - 2 >= 1:
            month = _month_on_row(ws, row - 2)
        year, month_from_dates = _first_date_in_section(ws, row + 1, starts)
        if month is None:
            month = month_from_dates
        if year is None and month_from_dates:
            year = year or _first_date_in_section(ws, row + 1, starts)[0]

        rr = row + 1
        found_summary = False
        while rr <= ws.max_row:
            if _is_shift_header(ws, rr):
                break
            summary = _summary_hours(ws, rr, starts, hours_cols)
            if summary:
                if year is None or month is None:
                    y, m = _first_date_in_section(ws, row + 1, starts)
                    year = year or y
                    month = month or m
                for start, hours in summary.items():
                    key = names.get(start)
                    if key and year and month:
                        totals[(year, month, key)] += hours
                found_summary = True
                rr += 1
                while rr <= ws.max_row:
                    fond = False
                    for hours_col in hours_cols:
                        label = ws.cell(rr, hours_col - 1).value
                        if isinstance(label, str) and label.strip().lower() == 'fond':
                            fond = True
                            break
                    if fond:
                        rr += 1
                        break
                    if _is_shift_header(ws, rr):
                        break
                    rr += 1
                break
            rr += 1

        if not found_summary and names:
            y, m = _first_date_in_section(ws, row + 1, starts)
            year = year or y
            month = month or m
            daily = defaultdict(float)
            rr = row + 1
            while rr <= ws.max_row and not _is_shift_header(ws, rr):
                for start, hours_col in zip(starts, hours_cols):
                    dt = ws.cell(rr, start).value
                    if not isinstance(dt, datetime):
                        continue
                    key = surname_key(ws.cell(rr, start + 1).value) or names.get(start)
                    h = _num(ws.cell(rr, hours_col).value)
                    if key and h is not None:
                        daily[(dt.year, dt.month, key)] += h
                rr += 1
            for key, hours in daily.items():
                totals[key] += hours

        row = rr

    return totals


def parse_smeny_files(paths):
    merged = defaultdict(float)
    for path in paths:
        for key, hours in parse_smeny_workbook(path).items():
            merged[key] += hours
    return merged


def build_override_payload(hours_map, rok=2026, mesice=(3, 4, 5)):
    users = defaultdict(list)
    for (year, month, surname), hours in sorted(hours_map.items()):
        if year != rok or month not in mesice:
            continue
        if hours <= 0:
            continue
        users[surname].append({
            'rok': year,
            'mesic': month,
            'odpracovano_h': round(float(hours), 2),
        })
    for surname in users:
        users[surname].sort(key=lambda r: r['mesic'])
    return {'uzivatele': dict(sorted(users.items()))}


def import_prumer_hodiny_from_excels(paths, rok=2026, mesice=(3, 4, 5)):
    merged = parse_smeny_files(paths)
    payload = build_override_payload(merged, rok=rok, mesice=mesice)
    return payload, merged


def write_prumer_override_json(payload, out_path):
    out_path = Path(out_path)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + '\n',
        encoding='utf-8',
    )
