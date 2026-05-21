from django.shortcuts import render
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from datetime import datetime, timedelta, date
import calendar
import csv
from .models import Smena, SmenaDochazka, SmenaStatistiky
from users.models import WebUser
from stores.models import Prodejna

def get_ceske_svatky(rok):
    """Vrací seznam českých státních svátků pro daný rok"""
    svatky = [
        (rok, 1, 1),    # Nový rok
        (rok, 5, 1),    # Svátek práce  
        (rok, 5, 8),    # Den vítězství
        (rok, 7, 5),    # Cyril a Metoděj
        (rok, 7, 6),    # Jan Hus
        (rok, 9, 28),   # Den české státnosti
        (rok, 10, 28),  # Vznik samostatného československého státu
        (rok, 11, 17),  # Den boje za svobodu a demokracii
        (rok, 12, 24),  # Štědrý den
        (rok, 12, 25),  # 1. svátek vánoční
        (rok, 12, 26),  # 2. svátek vánoční
    ]
    
    # Velikonoční svátky (závislé na datu Velikonoc)
    if rok == 2025:
        svatky.extend([
            (2025, 4, 18),  # Velký pátek
            (2025, 4, 21),  # Velikonoční pondělí
        ])
    elif rok == 2026:
        svatky.extend([
            (2026, 4, 3),   # Velký pátek
            (2026, 4, 6),   # Velikonoční pondělí
        ])
    # Pro další roky by bylo třeba dopočítat datum Velikonoc
    
    return svatky

def get_nazev_svatku(mesic, den):
    """Vrací název českého státního svátku podle data"""
    svatky_nazvy = {
        (1, 1): "Nový rok",
        (5, 1): "Svátek práce", 
        (5, 8): "Den vítězství",
        (7, 5): "Cyril a Metoděj",
        (7, 6): "Jan Hus", 
        (9, 28): "Den české státnosti",
        (10, 28): "Vznik samostatného československého státu",
        (11, 17): "Den boje za svobodu a demokracii",
        (12, 24): "Štědrý den",
        (12, 25): "1. svátek vánoční",
        (12, 26): "2. svátek vánoční",
        (4, 18): "Velký pátek",  # 2025
        (4, 21): "Velikonoční pondělí",  # 2025
        (4, 3): "Velký pátek",   # 2026  
        (4, 6): "Velikonoční pondělí",   # 2026
    }
    return svatky_nazvy.get((mesic, den), "Státní svátek")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def smeny_count(request):
    """
    GET – Počet pracovních směn prodejce na prodejně v daném měsíci.
    Query: user_id, prodejna_id, mesic (YYYY-MM).
    Pouze typ_smeny='prace', aktivni=True.
    """
    if request.user.role not in ['ADMIN', 'VEDOUCI']:
        return Response({'error': 'Přístup pouze pro vedoucí.'}, status=status.HTTP_403_FORBIDDEN)

    user_id = request.GET.get('user_id')
    prodejna_id = request.GET.get('prodejna_id')
    mesic_str = request.GET.get('mesic')  # YYYY-MM

    if not all([user_id, prodejna_id, mesic_str]):
        return Response({'error': 'Chybí parametry: user_id, prodejna_id, mesic'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rok, mesic_cislo = map(int, mesic_str.split('-'))
    except ValueError:
        return Response({'error': 'Neplatný formát měsíce (očekáváno YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)

    pocet = Smena.objects.filter(
        user_id=user_id,
        prodejna_id=prodejna_id,
        datum__year=rok,
        datum__month=mesic_cislo,
        typ_smeny='prace',
        aktivni=True,
    ).count()

    return Response({'pocet_smen': pocet})


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def smeny_list(request):
    """Seznam směn s filtrováním"""
    
    if request.method == 'GET':
        # Parametry pro filtrování
        mesic = request.GET.get('mesic')  # YYYY-MM
        prodejna = request.GET.get('prodejna')
        user_id = request.GET.get('user_id')
        
        smeny = Smena.objects.filter(aktivni=True)
        
        # Filtry pro prodejce - může vidět jen své směny a jen aktuální/budoucí měsíce
        if request.user.role not in ['ADMIN', 'VEDOUCI']:
            smeny = smeny.filter(user=request.user)
            # Omezení na aktuální a budoucí měsíce (plus minulý měsíc jen ke čtení)
            ted = date.today()
            minuly_mesic = ted.replace(day=1) - timedelta(days=1)
            minuly_mesic = minuly_mesic.replace(day=1)
            smeny = smeny.filter(datum__gte=minuly_mesic)
        
        # Filtry podle parametrů
        if mesic:
            try:
                rok, mesic_cislo = map(int, mesic.split('-'))
                smeny = smeny.filter(datum__year=rok, datum__month=mesic_cislo)
            except ValueError:
                pass
        
        if prodejna:
            smeny = smeny.filter(prodejna=prodejna)
            
        if user_id and request.user.role in ['ADMIN', 'VEDOUCI']:
            smeny = smeny.filter(user_id=user_id)
        
        # Načtení související dat
        smeny = smeny.select_related('user', 'prodejna').prefetch_related('dochazka').order_by('datum', 'cas_od')
        
        data = []
        for smena in smeny:
            data.append({
                'id': smena.id,
                'user_id': smena.user.id,
                'user_jmeno': smena.user.prijmeni,
                'user_prijmeni': smena.user.prijmeni,
                'prodejna_id': smena.prodejna.id,
                'prodejna_nazev': smena.prodejna.nazev,
                'prodejna': smena.prodejna.nazev_kratkiy,  # Použije krátký název pro frontend
                'user': {  # Pro kompatibilitu s frontendem
                    'jmeno': smena.user.jmeno,
                    'prijmeni': smena.user.prijmeni,
                },
                'datum': smena.datum,
                'cas_od': smena.cas_od,
                'cas_do': smena.cas_do,
                'typ_smeny': smena.typ_smeny,
                'poznamka': smena.poznamka,
                'je_domaci_prodejna': smena.je_domaci_prodejna,
                'delka_smeny_hodin': smena.delka_smeny_hodin,
                'dochazka': [
                    {
                        'id': d.id,
                        'typ_akce': d.typ_akce,
                        'cas': d.cas,
                        'poznamka': d.poznamka
                    } for d in smena.dochazka.all()
                ]
            })
        
        return Response(data)
    
    elif request.method == 'POST':
        # Vytvoření nové směny
        data = request.data
        
        
        # Kontrola oprávnění
        user_id = data.get('user_id', request.user.id)
        if request.user.role not in ['ADMIN', 'VEDOUCI'] and user_id != request.user.id:
            return Response({'error': 'Nemáte oprávnění vytvářet směny pro jiné uživatele'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        # Kontrola času - prodejci mohou editovat jen aktuální a budoucí měsíc
        datum_str = data.get('datum')
        if datum_str:
            datum = datetime.strptime(datum_str, '%Y-%m-%d').date()
            ted = date.today()
            aktualni_mesic = ted.replace(day=1)
            
            if request.user.role not in ['ADMIN', 'VEDOUCI'] and datum < aktualni_mesic:
                return Response({'error': 'Nelze vytvářet směny v minulých měsících'}, 
                              status=status.HTTP_403_FORBIDDEN)
        
        try:
            user = WebUser.objects.get(id=user_id)
            # Převod prodejny: přijímáme ID i název
            prodejna_input = data.get('prodejna')
            if prodejna_input is None:
                return Response({'error': 'Chybí parametr prodejna'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                # Číslo nebo string čísla
                prodejna_id = int(prodejna_input)
                prodejna_obj = Prodejna.objects.get(id=prodejna_id)
            except (ValueError, TypeError):
                # Hledání podle názvu (více polí)
                prodejna_obj = Prodejna.objects.filter(
                    Q(nazev__iexact=prodejna_input) |
                    Q(nazev_kratkiy__iexact=prodejna_input) |
                    Q(nazev_google_sheets__iexact=prodejna_input)
                ).first()
                if not prodejna_obj:
                    return Response({'error': f"Prodejna '{prodejna_input}' nebyla nalezena"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Kontrola, zda směna už neexistuje
            existing_smena = Smena.objects.filter(
                user=user,
                datum=data['datum'],
                prodejna=prodejna_obj,
                aktivni=True
            ).first()
            
            if existing_smena:
                return Response({
                    'error': f'Na datum {data["datum"]} už máte směnu v prodejně {data["prodejna"]}. Chcete-li změnit čas, upravte stávající směnu.',
                    'existing_shift_id': existing_smena.id,
                    'existing_shift': {
                        'cas_od': existing_smena.cas_od.strftime('%H:%M'),
                        'cas_do': existing_smena.cas_do.strftime('%H:%M'),
                        'typ_smeny': existing_smena.typ_smeny
                    }
                }, status=status.HTTP_409_CONFLICT)
            
            smena = Smena.objects.create(
                user=user,
                prodejna=prodejna_obj,
                datum=data['datum'],
                cas_od=data['cas_od'],
                cas_do=data['cas_do'],
                typ_smeny=data.get('typ_smeny', 'prace'),
                poznamka=data.get('poznamka', '')
            )
            
            return Response({
                'id': smena.id,
                'message': 'Směna byla úspěšně vytvořena'
            }, status=status.HTTP_201_CREATED)
            
        except WebUser.DoesNotExist:
            return Response({'error': 'Uživatel neexistuje'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def smeny_bulk_create(request):
    """Hromadné vytvoření směn"""
    
    data = request.data
    user_id = data.get('user_id', request.user.id)
    
    # Kontrola oprávnění
    if request.user.role not in ['ADMIN', 'VEDOUCI'] and user_id != request.user.id:
        return Response({'error': 'Nemáte oprávnění vytvářet směny pro jiné uživatele'}, 
                      status=status.HTTP_403_FORBIDDEN)
    
    try:
        user = WebUser.objects.get(id=user_id)
        datumy = data.get('datumy', [])  # Seznam datumů
        prodejna_input = data.get('prodejna')
        cas_od = data['cas_od']
        cas_do = data['cas_do']
        typ_smeny = data.get('typ_smeny', 'prace')
        poznamka = data.get('poznamka', '')
        
        
        # Kontrola času pro prodejce
        ted = date.today()
        aktualni_mesic = ted.replace(day=1)
        
        uspesne = 0
        chyby = []
        
        # Mapování prodejny (ID nebo název)
        if prodejna_input is None:
            return Response({'error': 'Chybí parametr prodejna'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            prodejna_id = int(prodejna_input)
            prodejna_obj = Prodejna.objects.get(id=prodejna_id)
        except (ValueError, TypeError):
            prodejna_obj = Prodejna.objects.filter(
                Q(nazev__iexact=prodejna_input) |
                Q(nazev_kratkiy__iexact=prodejna_input) |
                Q(nazev_google_sheets__iexact=prodejna_input)
            ).first()
            if not prodejna_obj:
                return Response({'error': f"Prodejna '{prodejna_input}' nebyla nalezena"}, status=status.HTTP_400_BAD_REQUEST)

        for datum_str in datumy:
            try:
                datum = datetime.strptime(datum_str, '%Y-%m-%d').date()
                
                # Kontrola oprávnění pro datum
                if request.user.role not in ['ADMIN', 'VEDOUCI'] and datum < aktualni_mesic:
                    chyby.append(f'{datum_str}: Nelze vytvářet směny v minulých měsících')
                    continue
                
                # Kontrola, zda směna již neexistuje
                if Smena.objects.filter(user=user, datum=datum, prodejna=prodejna_obj, aktivni=True).exists():
                    chyby.append(f'{datum_str}: Směna již existuje')
                    continue
                
                Smena.objects.create(
                    user=user,
                    prodejna=prodejna_obj,
                    datum=datum,
                    cas_od=cas_od,
                    cas_do=cas_do,
                    typ_smeny=typ_smeny,
                    poznamka=poznamka
                )
                uspesne += 1
                
            except Exception as e:
                chyby.append(f'{datum_str}: {str(e)}')
        
        return Response({
            'uspesne': uspesne,
            'chyby': chyby,
            'message': f'Úspěšně vytvořeno {uspesne} směn'
        })
        
    except WebUser.DoesNotExist:
        return Response({'error': 'Uživatel neexistuje'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def smena_detail(request, smena_id):
    """Úprava a mazání směny"""
    
    smena = get_object_or_404(Smena, id=smena_id, aktivni=True)
    
    # Kontrola oprávnění
    if request.user.role not in ['ADMIN', 'VEDOUCI'] and smena.user != request.user:
        return Response({'error': 'Nemáte oprávnění upravovat tuto směnu'}, 
                      status=status.HTTP_403_FORBIDDEN)
    
    # Kontrola času pro prodejce
    if request.user.role not in ['ADMIN', 'VEDOUCI']:
        ted = date.today()
        aktualni_mesic = ted.replace(day=1)
        if smena.datum < aktualni_mesic:
            return Response({'error': 'Nelze upravovat směny v minulých měsících'}, 
                          status=status.HTTP_403_FORBIDDEN)
    
    if request.method == 'PUT':
        try:
            data = request.data
            
            # Aktualizace dat
            if 'prodejna' in data:
                smena.prodejna = data['prodejna']
            if 'datum' in data:
                smena.datum = data['datum']
            if 'cas_od' in data:
                smena.cas_od = data['cas_od']
            if 'cas_do' in data:
                smena.cas_do = data['cas_do']
            if 'typ_smeny' in data:
                smena.typ_smeny = data['typ_smeny']
            if 'poznamka' in data:
                smena.poznamka = data['poznamka']
            
            smena.save()
            
            return Response({'message': 'Směna byla úspěšně aktualizována'})
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        # Všichni uživatelé (včetně prodejců) skutečně mažou z databáze
        smena_info = f"{smena.user.prijmeni} - {smena.datum} - {smena.prodejna}"
        smena.delete()
        return Response({'message': f'Směna byla úspěšně smazána z databáze: {smena_info}'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def kalendar_data(request):
    """Data pro kalendářní pohled"""
    
    mesic = request.GET.get('mesic')  # YYYY-MM
    prodejna_id = request.GET.get('prodejna')
    
    print(f"Kalendář požadavek - mesic: '{mesic}', prodejna_id: '{prodejna_id}'")
    print(f"Všechny GET parametry: {request.GET}")
    
    if not mesic or not prodejna_id:
        error_msg = f"Chybí parametry: mesic='{mesic}', prodejna_id='{prodejna_id}'"
        print(f"Chyba 400: {error_msg}")
        return Response({'error': error_msg}, 
                      status=status.HTTP_400_BAD_REQUEST)
    
    try:
        rok, mesic_cislo = map(int, mesic.split('-'))
        print(f"Parsované datum - rok: {rok}, měsíc: {mesic_cislo}")
        
        # Najdeme prodejnu podle ID
        try:
            from stores.models import Prodejna
            prodejna = Prodejna.objects.get(id=prodejna_id, aktivni=True)
            print(f"Našel jsem prodejnu: {prodejna.nazev} (ID: {prodejna.id})")
        except Prodejna.DoesNotExist:
            error_msg = f"Prodejna s ID '{prodejna_id}' nebyla nalezena nebo není aktivní"
            print(f"Chyba: {error_msg}")
            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            error_msg = f"Neplatné ID prodejny: '{prodejna_id}' - očekává se číslo"
            print(f"Chyba: {error_msg}")
            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
        
        # Získání směn pro daný měsíc a prodejnu
        smeny = Smena.objects.filter(
            datum__year=rok,
            datum__month=mesic_cislo,
            prodejna=prodejna,  # Použijeme objekt prodejny
            aktivni=True
        ).select_related('user').order_by('datum', 'cas_od')
        
        print(f"Načteno {smeny.count()} směn pro prodejnu '{prodejna.nazev}' v měsíci {mesic}")
        
        # Seskupení podle datumu
        kalendar_data = {}
        for smena in smeny:
            datum_str = smena.datum.strftime('%Y-%m-%d')
            if datum_str not in kalendar_data:
                kalendar_data[datum_str] = []
            
            kalendar_data[datum_str].append({
                'id': smena.id,
                'user_id': smena.user.id,
                'user_jmeno': smena.user.prijmeni,
                'cas_od': smena.cas_od.strftime('%H:%M'),
                'cas_do': smena.cas_do.strftime('%H:%M'),
                'typ_smeny': smena.typ_smeny,
                'je_domaci_prodejna': smena.je_domaci_prodejna
            })
        
        # Dnešní a zítřejší směny pro info boxy
        dnes = date.today()
        zitra = dnes + timedelta(days=1)
        
        dnes_smeny = smeny.filter(datum=dnes)
        zitra_smeny = smeny.filter(datum=zitra)
        
        # Státní svátky pro daný měsíc
        ceske_svatky = get_ceske_svatky(rok)
        svatky_mesic = {}
        for rok_s, mesic_s, den_s in ceske_svatky:
            if mesic_s == mesic_cislo:
                datum_str = f"{rok_s}-{mesic_s:02d}-{den_s:02d}"
                svatky_mesic[datum_str] = {
                    'je_svatek': True,
                    'nazev': get_nazev_svatku(mesic_s, den_s)
                }
        
        response_data = {
            'kalendar_data': kalendar_data,
            'svatky': svatky_mesic,
            'dnes_smeny': [
                f"{s.user.prijmeni} ({s.cas_od.strftime('%H:%M')}-{s.cas_do.strftime('%H:%M')})" 
                for s in dnes_smeny
            ],
            'zitra_smeny': [
                f"{s.user.prijmeni} ({s.cas_od.strftime('%H:%M')}-{s.cas_do.strftime('%H:%M')})" 
                for s in zitra_smeny
            ]
        }
        
        print(f"Úspěšně vrácena data pro kalendář - {len(kalendar_data)} dní se směnami")
        return Response(response_data)
        
    except ValueError as e:
        error_msg = f"Neplatný formát měsíce: {mesic} - {str(e)}"
        print(f"Chyba ValueError: {error_msg}")
        return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        error_msg = f"Neočekávaná chyba: {str(e)}"
        print(f"Chyba Exception: {error_msg}")
        return Response({'error': error_msg}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def dochazka_akce(request):
    """Check-in/check-out/pauza akce"""
    
    data = request.data
    smena_id = data.get('smena_id')
    typ_akce = data.get('typ_akce')
    poznamka = data.get('poznamka', '')
    
    if not smena_id or not typ_akce:
        return Response({'error': 'Chybí povinné parametry'}, 
                      status=status.HTTP_400_BAD_REQUEST)
    
    smena = get_object_or_404(Smena, id=smena_id, aktivni=True)
    
    # Kontrola oprávnění
    if request.user.role not in ['ADMIN', 'VEDOUCI'] and smena.user != request.user:
        return Response({'error': 'Nemáte oprávnění k této směně'}, 
                      status=status.HTTP_403_FORBIDDEN)
    
    try:
        dochazka = SmenaDochazka.objects.create(
            smena=smena,
            typ_akce=typ_akce,
            cas=datetime.now(),
            poznamka=poznamka
        )
        
        return Response({
            'id': dochazka.id,
            'message': f'{dochazka.get_typ_akce_display()} byl úspěšně zaznamenán',
            'cas': dochazka.cas
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def smeny_prehled(request):
    """Přehled hodin pro uživatele"""
    
    user_id = request.GET.get('user_id')
    mesic = request.GET.get('mesic')  # YYYY-MM
    
    # Kontrola oprávnění
    if user_id:
        if request.user.role != 'ADMIN' and int(user_id) != request.user.id:
            return Response({'error': 'Nemáte oprávnění'}, status=status.HTTP_403_FORBIDDEN)
        user = get_object_or_404(WebUser, id=user_id)
    else:
        user = request.user
    
    # Výchozí měsíc je aktuální
    if mesic:
        try:
            rok, mesic_cislo = map(int, mesic.split('-'))
        except ValueError:
            return Response({'error': 'Neplatný formát měsíce'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        dnes = date.today()
        rok, mesic_cislo = dnes.year, dnes.month
    
    # Získání směn pro daný měsíc
    smeny = Smena.objects.filter(
        user=user,
        datum__year=rok,
        datum__month=mesic_cislo,
        aktivni=True
    ).select_related('prodejna')
    
    # Výpočet statistik
    celkem_hodin = sum(smena.delka_smeny_hodin for smena in smeny if smena.typ_smeny == 'prace')
    hodin_dovolene = sum(smena.delka_smeny_hodin for smena in smeny if smena.typ_smeny == 'dovolena')
    pocet_smeny = smeny.filter(typ_smeny='prace').count()
    
    # Výpočet skutečných pracovních dnů v měsíci
    ceske_svatky = get_ceske_svatky(rok)
    pocet_dni_mesic = calendar.monthrange(rok, mesic_cislo)[1]
    pracovni_dny = 0
    
    for den in range(1, pocet_dni_mesic + 1):
        datum_kontrola = date(rok, mesic_cislo, den)
        # Není sobota ani neděle
        if datum_kontrola.weekday() < 5:  # 0-4 = Po-Pá
            # Není státní svátek
            if (rok, mesic_cislo, den) not in ceske_svatky:
                pracovni_dny += 1
    
    from .labor_hours import fondu_hodin_mesic
    mesicni_fond = fondu_hodin_mesic(rok, mesic_cislo)
    
    return Response({
        'mesic': f'{rok}-{mesic_cislo:02d}',
        'user_jmeno': user.prijmeni,
        'celkem_hodin_naplanovanych': celkem_hodin,
        'hodin_dovolene': hodin_dovolene,
        'pocet_smeny': pocet_smeny,
        'standardni_hodiny': mesicni_fond,
        'mesicni_fond': mesicni_fond,
        'procento_naplneni': round((celkem_hodin / mesicni_fond) * 100, 1) if mesicni_fond > 0 else 0,
        'smeny_detail': [
            {
                'datum': smena.datum,
                'prodejna': smena.prodejna.nazev,
                'cas_od': smena.cas_od,
                'cas_do': smena.cas_do,
                'hodiny': smena.delka_smeny_hodin,
                'typ_smeny': smena.typ_smeny,
                'je_domaci_prodejna': smena.je_domaci_prodejna
            } for smena in smeny.order_by('datum')
        ]
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_smeny(request):
    """Export směn a odměňování – pouze ADMIN, vše v bodech."""
    from io import BytesIO
    from .payroll_service import build_payroll_preview

    if request.user.role != 'ADMIN':
        return Response({'error': 'Nemáte oprávnění'}, status=status.HTTP_403_FORBIDDEN)

    mesic = request.GET.get('mesic')
    prodejna = request.GET.get('prodejna')

    if not mesic:
        return Response({'error': 'Chybí parametr mesic'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rok, mesic_cislo = map(int, mesic.split('-'))
        nazvy_mesicu = {
            1: 'Leden', 2: 'Únor', 3: 'Březen', 4: 'Duben',
            5: 'Květen', 6: 'Červen', 7: 'Červenec', 8: 'Srpen',
            9: 'Září', 10: 'Říjen', 11: 'Listopad', 12: 'Prosinec',
        }
        nazev_mesice = f"{nazvy_mesicu[mesic_cislo]} {rok}"

        preview = build_payroll_preview(mesic, prodejna_id=prodejna)
        rows = sorted(preview.get('rows') or [], key=lambda r: r.get('jmeno') or '')

        provize_cols = [
            ('polozky_nad_100', 'Položky nad 100'),
            ('ct600', 'CT600'),
            ('ct1200', 'CT1200'),
            ('akt', 'AKT'),
            ('zah250', 'ZAH250'),
            ('nap', 'NAP'),
            ('zah500', 'ZAH500'),
            ('kop250', 'KOP250'),
            ('kop500', 'KOP500'),
            ('pz1', 'PZ1'),
            ('knz', 'KNZ'),
            ('servis_marze', 'Servis'),
        ]
        doplnek_kody = [
            ('vedouci_pobocky', 'Vedoucí pobočky'),
        ]

        headers = [
            'Měsíc', 'Jméno', 'Středisko',
            'Odpracováno h', 'Dovolená h', 'Nemoc h', 'Svátek h',
            'Fond h', 'Přesčas h',
            'Základ', 'Doplňky',
        ]
        headers += [label for _k, label in doplnek_kody]
        headers += [label for _k, label in provize_cols]
        headers += ['Provize celkem', 'Odměna měsíc', 'Celkem']

        def row_values(data):
            bd = data.get('provize_breakdown') or {}
            doplnky_by_kod = {d.get('kod'): d.get('castka', 0) for d in (data.get('doplnky') or [])}
            out = [
                nazev_mesice,
                data.get('jmeno'),
                data.get('stredisko'),
                data.get('odpracovano_h', 0),
                data.get('dovolena_h', 0),
                data.get('nemoc_h', 0),
                data.get('svatek_h', 0),
                data.get('fondu_h', 0),
                data.get('prescas_h', 0),
                data.get('zaklad_body', 0),
                data.get('doplnky_body', 0),
            ]
            for kod, _label in doplnek_kody:
                out.append(doplnky_by_kod.get(kod, 0))
            for key, _label in provize_cols:
                item = bd.get(key) or {}
                out.append(item.get('points', 0))
            out.extend([
                data.get('provize_body', 0),
                data.get('odmena_mesic_body', 0),
                data.get('celkem_body', 0),
            ])
            return out

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = f"Výplata {mesic}"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row_idx = 2
            for data in rows:
                for col, val in enumerate(row_values(data), 1):
                    ws.cell(row=row_idx, column=col, value=val).border = thin_border
                row_idx += 1

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="vyplata_{mesic}.xlsx"'
            return response

        except ImportError:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="vyplata_{mesic}.csv"'
            response.write('\ufeff')
            writer = csv.writer(response, delimiter=';')
            writer.writerow(headers)
            for data in rows:
                writer.writerow(row_values(data))
            return response

    except ValueError as e:
        return Response({'error': f'Neplatný formát měsíce: {e}'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'Chyba při exportu: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

