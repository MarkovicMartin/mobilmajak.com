"""Import jednotlivých směn z Excelů Směny <prodejna>.xlsx do tabulky WEB_SMENY."""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

from django.db import transaction

from shifts.models import Smena
from shifts.shift_helpers import find_overlapping_shift, is_absence_shift, is_backoffice_user, resolve_prodejna
from shifts.vacation_service import normalize_dovolena_casy
from stores.models import Prodejna
from stores.oteviraci_doba_utils import DNY_KLICE, resolve_den_hours
from users.models import WebUser

from .smeny_excel_import import (
    ALIASES,
    MONTH_NAMES,
    SKIP_TOKENS,
    _first_date_in_section,
    _hours_col_for_block,
    _hours_from_times,
    _is_shift_header,
    _month_on_row,
    _names_for_blocks,
    _num,
    _shift_header_starts,
    surname_key,
)

IMPORT_POZNAMKA_PREFIX = 'Excel import'

STORE_ALIASES = {
    'globus': 'Globus',
    'servis gl': 'Globus',
    'servis globus': 'Globus',
    'senimo': 'Senimo',
    'hlavní sklad': 'Senimo',
    'hlavni sklad': 'Senimo',
    'zlín': 'Zlín',
    'zlin': 'Zlín',
    'čepkov': 'Zlín',
    'cepkov': 'Zlín',
    'přerov': 'Přerov',
    'prerov': 'Přerov',
    'vsetín': 'Vsetín',
    'vsetin': 'Vsetín',
    'šternberk': 'Šternberk',
    'sternberk': 'Šternberk',
}

ABSENT_ARRIVAL = {'x', 'xx', '-', '—'}

BRIGADNIK_NAME_ALIASES = {
    'monika': 'křížková',
    'kuba': 'málek',
}

BRIGADNIK_SHEET_HINTS = ('brigádn', 'brigadn')


@dataclass
class ParsedShiftRow:
    datum: date
    surname_key: str
    raw_name: str
    prodejna_nazev: str
    cas_od: time | None
    cas_do: time | None
    hours: float | None
    typ_smeny: str
    pozice_smeny: str
    brigadnik_rezim: str
    poznamka_extra: str
    source_file: str
    times_from_excel: bool = False
    hours_from_excel: bool = False
    from_brigadnik_sheet: bool = False


def shift_surname_key(name):
    if name is None:
        return None
    raw = str(name).strip().lstrip('*')
    if not raw:
        return None
    cleaned = re.sub(
        r'(?i)\b(nemocensk[áa]|dovolen[áa]|dovol)\b',
        '',
        raw,
    ).strip(' -–')
    key = surname_key(cleaned) or surname_key(raw)
    if key:
        return BRIGADNIK_NAME_ALIASES.get(key, key)
    return None


def default_store_from_path(path: Path | str) -> str | None:
    stem = Path(path).stem.lower()
    for needle, store in [
        ('šternberk', 'Šternberk'),
        ('sternberk', 'Šternberk'),
        ('přerov', 'Přerov'),
        ('prerov', 'Přerov'),
        ('zlín', 'Zlín'),
        ('zlin', 'Zlín'),
        ('senimo', 'Senimo'),
        ('globus', 'Globus'),
        ('vsetín', 'Vsetín'),
        ('vsetin', 'Vsetín'),
    ]:
        if needle in stem:
            return store
    return None


def resolve_store_label(text, default_store: str | None) -> tuple[str | None, str]:
    """Vrátí (název prodejny, poznámka navíc)."""
    if not text:
        return default_store, ''
    raw = str(text).strip()
    low = raw.lower()
    if not raw or low in SKIP_TOKENS:
        return default_store, ''
    for alias, store in STORE_ALIASES.items():
        if alias in low:
            extra = raw if store.lower() not in low else ''
            return store, extra.strip()
    if default_store:
        return default_store, raw
    return None, raw


def detect_shift_type(name, store_text, hours) -> str:
    blob = f'{name or ""} {store_text or ""}'.lower()
    if 'nemoc' in blob:
        return 'nemoc'
    if 'dovo' in blob or 'dovol' in blob:
        return 'dovolena'
    if hours is not None and hours <= 0 and 'svátek' not in blob and 'svatek' not in blob:
        return 'nemoc' if 'nemoc' in blob else 'prace'
    return 'prace'


def parse_excel_time_cell(value):
    if value is None:
        return None
    if isinstance(value, str):
        low = value.strip().lower()
        if low in ABSENT_ARRIVAL or low in SKIP_TOKENS:
            return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        whole = int(value)
        minutes = int(round((float(value) - whole) * 60))
        if minutes >= 60:
            whole += minutes // 60
            minutes = minutes % 60
        return time(whole % 24, minutes)
    if isinstance(value, str):
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def _block_layout(ws, header_row, start):
    hours_col = _hours_col_for_block(ws, header_row, start)
    name_col = start + 1
    store_col = None
    od_col = None
    do_col = None
    for off in range(0, 8):
        col = start + off
        if col > ws.max_column:
            break
        val = ws.cell(header_row, col).value
        if not isinstance(val, str):
            continue
        low = val.strip().lower()
        if low.startswith('jméno') or low.startswith('jmeno'):
            name_col = col
        elif 'prodej' in low:
            store_col = col
        elif 'příchod' in low or 'prichod' in low:
            od_col = col
        elif low == 'odchod':
            do_col = col
    if store_col is None:
        store_col = name_col + 1
    return {
        'start': start,
        'name_col': name_col,
        'store_col': store_col,
        'od_col': od_col,
        'do_col': do_col,
        'hours_col': hours_col,
    }


def opening_hours_for_store(store, shift_date: date) -> tuple[time, time, float] | None:
    if not store:
        return None
    den_key = DNY_KLICE[shift_date.weekday()]
    pair = resolve_den_hours(store.oteviraci_doba, den_key)
    if not pair:
        return None
    cas_od = datetime.strptime(pair[0], '%H:%M').time()
    cas_do = datetime.strptime(pair[1], '%H:%M').time()
    hours = _hours_from_times(cas_od, cas_do)
    if hours is None or hours <= 0:
        return None
    return cas_od, cas_do, hours


def apply_opening_hours_fallback(shifts, store_lookup=None):
    """Chybí-li hodiny/časy a na prodejně je jen jeden člověk → otevírací doba."""
    store_lookup = store_lookup or build_prodejna_lookup()
    by_day_store = defaultdict(list)
    for idx, shift in enumerate(shifts):
        if shift.typ_smeny != 'prace' or not shift.prodejna_nazev:
            continue
        by_day_store[(shift.datum, shift.prodejna_nazev.lower())].append(idx)

    for (_datum, store_name), indices in by_day_store.items():
        if len(indices) != 1:
            continue
        shift = shifts[indices[0]]
        store = store_lookup.get(store_name)
        opening = opening_hours_for_store(store, shift.datum)
        if not opening:
            continue
        cas_od, cas_do, open_hours = opening
        if not shift.hours_from_excel:
            shift.cas_od = cas_od
            shift.cas_do = cas_do
            shift.hours = open_hours
        elif not shift.times_from_excel and shift.hours_from_excel:
            shift.cas_od = cas_od
            end_dt = datetime.combine(date.min, cas_od) + timedelta(hours=float(shift.hours))
            shift.cas_do = end_dt.time()


def _shift_row_rank(shift: ParsedShiftRow) -> tuple:
    return (
        0 if shift.from_brigadnik_sheet else 1,
        1 if shift.times_from_excel else 0,
        1 if shift.hours_from_excel else 0,
        float(shift.hours or 0),
    )


def deduplicate_parsed_shifts(shifts):
    """Stejný den/uživatel/prodejna z více Excelů → nejúplnější záznam."""
    best = {}
    for shift in shifts:
        key = (shift.surname_key, shift.datum, shift.prodejna_nazev.lower(), shift.typ_smeny)
        prev = best.get(key)
        if prev is None or _shift_row_rank(shift) > _shift_row_rank(prev):
            best[key] = shift
    return list(best.values())


def infer_shift_times(cas_od, cas_do, hours, typ_smeny) -> tuple[time, time]:
    if is_absence_shift(typ_smeny):
        return time(8, 0), time(16, 0)

    if cas_od is not None and cas_do is not None:
        return cas_od, cas_do

    if cas_od is not None and hours is not None and hours > 0:
        start_dt = datetime.combine(date.min, cas_od)
        end_dt = start_dt + timedelta(hours=float(hours))
        return cas_od, end_dt.time()

    if hours is not None and hours > 0:
        start_dt = datetime.combine(date.min, time(8, 0))
        end_dt = start_dt + timedelta(hours=float(hours))
        return time(8, 0), end_dt.time()

    return time(8, 0), time(16, 0)


def _row_shift_hours(ws, row, layout):
    hours_col = layout['hours_col']
    h = _num(ws.cell(row, hours_col).value)
    if h is not None:
        return max(0.0, h)
    od_col = layout['od_col']
    do_col = layout['do_col']
    if od_col and do_col:
        od = ws.cell(row, od_col).value
        do = ws.cell(row, do_col).value
        if isinstance(od, (int, float)) and isinstance(do, (int, float)):
            return max(0.0, float(do) - float(od))
        computed = _hours_from_times(parse_excel_time_cell(od), parse_excel_time_cell(do))
        if computed is not None:
            return max(0.0, computed)
    return None


def _is_brigadnik_sheet_name(name: str) -> bool:
    low = (name or '').lower()
    return any(h in low for h in BRIGADNIK_SHEET_HINTS)


def _workbook_sheets_to_parse(wb):
    """Vrátí (název listu, je_brigádníci) pro všechny relevantní listy."""
    sheets = []
    for sn in wb.sheetnames:
        low = sn.lower()
        if low.startswith('list ') and low not in ('list 1',):
            continue
        if _is_brigadnik_sheet_name(sn):
            sheets.append((sn, True))
        elif low in ('list 1', 'zaměstnanci', 'zamestnanci') or len(wb.sheetnames) == 1:
            sheets.append((sn, False))
    if not sheets:
        sheets.append((wb.active.title, False))
    return sheets


def _is_compact_shift_header(ws, row):
    v1 = ws.cell(row, 1).value
    v2 = ws.cell(row, 2).value
    if not isinstance(v1, str) or v1.strip().lower() != 'den':
        return False
    if not isinstance(v2, str) or not v2.strip().lower().startswith('jm'):
        return False
    for col in range(1, min(ws.max_column, 24) + 1):
        val = ws.cell(row, col).value
        if isinstance(val, str) and 'prodej' in val.strip().lower():
            return False
        if isinstance(val, str) and val.strip().lower() == 'den' and col > 2:
            return False
    v6 = ws.cell(row, 6).value
    return isinstance(v6, str) and (
        'hod' in v6.strip().lower() or 'počet' in v6.strip().lower() or 'pocet' in v6.strip().lower()
    )


def _compact_layout():
    return {
        'start': 1,
        'name_col': 2,
        'store_col': 5,
        'od_col': 3,
        'do_col': 4,
        'hours_col': 6,
    }


def _is_summary_row(ws, row, layouts):
    label_cells = [
        ws.cell(row, layout['hours_col'] - 1).value
        for layout in layouts
        if layout['hours_col'] > 1
    ]
    return any(
        isinstance(v, str) and v.strip().lower() in {
            'celkem', 'součet', 'soucet', 'fond', 'rozdíl', 'rozdil',
        }
        for v in label_cells
    )


def _parse_row_block(
    ws, row, layout, default_store, source_file, rok, mesice,
    *, from_brigadnik_sheet=False,
):
    start = layout['start']
    dt_val = ws.cell(row, start).value
    if not isinstance(dt_val, datetime):
        return None
    shift_date = dt_val.date()
    if shift_date.year != rok or shift_date.month not in mesice:
        return None

    raw_name = ws.cell(row, layout['name_col']).value
    key = shift_surname_key(raw_name)
    if not key:
        return None

    store_text = ws.cell(row, layout['store_col']).value if layout['store_col'] else None
    prodejna_nazev, extra_note = resolve_store_label(store_text, default_store)

    od_val = ws.cell(row, layout['od_col']).value if layout['od_col'] else None
    do_val = ws.cell(row, layout['do_col']).value if layout['do_col'] else None
    if isinstance(od_val, str) and od_val.strip().lower() in ABSENT_ARRIVAL:
        return None

    hours = _row_shift_hours(ws, row, layout)
    cas_od = parse_excel_time_cell(od_val)
    cas_do = parse_excel_time_cell(do_val)
    times_from_excel = cas_od is not None or cas_do is not None
    if (hours is None or hours <= 0) and times_from_excel:
        inferred = _hours_from_times(cas_od, cas_do)
        if inferred is not None and inferred > 0:
            hours = inferred

    typ_smeny = detect_shift_type(raw_name, store_text, hours)
    hours_from_excel = hours is not None and hours > 0

    if typ_smeny == 'prace' and not hours_from_excel and not times_from_excel:
        if not prodejna_nazev:
            return None
    elif typ_smeny == 'prace' and (hours is None or hours <= 0):
        return None

    cas_od, cas_do = infer_shift_times(cas_od, cas_do, hours, typ_smeny)

    pozice = 'servis' if store_text and 'servis' in str(store_text).lower() else 'prodej'
    brigadnik_rezim = 'vypomoc' if (
        from_brigadnik_sheet
        or (store_text and 'výpomoc' in str(store_text).lower())
        or (
            prodejna_nazev
            and default_store
            and prodejna_nazev.lower() != default_store.lower()
        )
    ) else 'prodejce'
    if not prodejna_nazev and typ_smeny == 'prace':
        return None

    return ParsedShiftRow(
        datum=shift_date,
        surname_key=key,
        raw_name=str(raw_name).strip(),
        prodejna_nazev=prodejna_nazev or '',
        cas_od=cas_od,
        cas_do=cas_do,
        hours=hours,
        typ_smeny=typ_smeny,
        pozice_smeny=pozice,
        brigadnik_rezim=brigadnik_rezim,
        poznamka_extra=extra_note,
        source_file=source_file,
        times_from_excel=times_from_excel,
        hours_from_excel=hours_from_excel,
        from_brigadnik_sheet=from_brigadnik_sheet,
    )


def parse_shifts_worksheet(
    ws, default_store, source_file, rok, mesice, *, from_brigadnik_sheet=False,
):
    shifts = []
    row = 1
    while row <= ws.max_row:
        if _is_compact_shift_header(ws, row):
            layout = _compact_layout()
            rr = row + 1
            while rr <= ws.max_row:
                if _is_compact_shift_header(ws, rr) or _is_shift_header(ws, rr):
                    break
                if _is_summary_row(ws, rr, [layout]):
                    rr += 1
                    continue
                parsed = _parse_row_block(
                    ws, rr, layout, default_store, source_file, rok, mesice,
                    from_brigadnik_sheet=from_brigadnik_sheet,
                )
                if parsed:
                    shifts.append(parsed)
                rr += 1
            row = rr
            continue

        if not _is_shift_header(ws, row):
            row += 1
            continue

        starts = _shift_header_starts(ws, row)
        if not starts:
            row += 1
            continue

        layouts = [_block_layout(ws, row, start) for start in starts]
        rr = row + 1
        while rr <= ws.max_row:
            if _is_shift_header(ws, rr) or _is_compact_shift_header(ws, rr):
                break
            if _is_summary_row(ws, rr, layouts):
                rr += 1
                continue

            for layout in layouts:
                parsed = _parse_row_block(
                    ws, rr, layout, default_store, source_file, rok, mesice,
                    from_brigadnik_sheet=from_brigadnik_sheet,
                )
                if parsed:
                    shifts.append(parsed)
            rr += 1
        row = rr

    return shifts


def parse_shifts_workbook(path, rok=2026, mesice=(3, 4, 5)):
    import openpyxl

    path = Path(path)
    default_store = default_store_from_path(path)
    source_file = path.name
    mesice = tuple(mesice)
    shifts = []

    wb = openpyxl.load_workbook(path, data_only=True)
    shifts = []
    for sheet_name, is_brigadnik in _workbook_sheets_to_parse(wb):
        ws = wb[sheet_name]
        shifts.extend(parse_shifts_worksheet(
            ws, default_store, source_file, rok, mesice,
            from_brigadnik_sheet=is_brigadnik,
        ))

    apply_opening_hours_fallback(shifts)
    shifts = [
        s for s in shifts
        if s.typ_smeny != 'prace' or (s.hours is not None and s.hours > 0)
    ]
    return shifts


def parse_shifts_files(paths, rok=2026, mesice=(3, 4, 5)):
    all_shifts = []
    for path in paths:
        all_shifts.extend(parse_shifts_workbook(path, rok=rok, mesice=mesice))
    return deduplicate_parsed_shifts(all_shifts)


def filter_brigadnik_only_users(shifts, user_lookup):
    """List Brigádníci je jen pro roli BRIGADNIK – ostatní už mají směny na hlavním listu."""
    kept = []
    for shift in shifts:
        if not shift.from_brigadnik_sheet:
            kept.append(shift)
            continue
        user = user_lookup.get(shift.surname_key)
        if user and getattr(user, 'role', None) != 'BRIGADNIK':
            continue
        kept.append(shift)
    return kept


def build_user_lookup():
    lookup = {}
    duplicates = {}
    skip_prijmeni = {'prodejce', 'nový', 'novy', 'admin', 'test'}
    role_rank = {'PRODEJCE': 0, 'VEDOUCI': 1, 'BRIGADNIK': 2, 'ADMIN': 3}

    def score(user):
        return (
            0 if user.aktivni else 1,
            role_rank.get(user.role, 9),
            user.id,
        )

    for user in sorted(WebUser.objects.all(), key=score):
        prijmeni_low = (user.prijmeni or '').strip().lower()
        if prijmeni_low in skip_prijmeni:
            continue
        keys = set()
        if user.prijmeni:
            keys.add(surname_key(user.prijmeni) or prijmeni_low)
        full = f'{user.jmeno} {user.prijmeni}'.strip()
        if full:
            keys.add(surname_key(full) or full.strip().lower())
        for key in keys:
            if not key or key in skip_prijmeni:
                continue
            if key in lookup and lookup[key].id != user.id:
                duplicates.setdefault(key, set()).update({lookup[key].id, user.id})
            else:
                lookup[key] = user
    for alias, target in ALIASES.items():
        if target in lookup and alias not in lookup:
            lookup[alias] = lookup[target]
    return lookup, duplicates


def build_prodejna_lookup():
    stores = {}
    for store in Prodejna.objects.filter(aktivni=True):
        for label in (store.nazev, store.nazev_kratkiy, store.nazev_google_sheets):
            if label:
                stores[label.strip().lower()] = store
    for alias, nazev in STORE_ALIASES.items():
        if nazev.lower() not in stores:
            obj = Prodejna.objects.filter(nazev=nazev).first()
            if obj:
                stores[nazev.lower()] = obj
                stores[alias] = obj
    return stores


def import_poznamka(rok, mesice):
    mesice_txt = ','.join(f'{m:02d}' for m in sorted(mesice))
    return f'{IMPORT_POZNAMKA_PREFIX} {rok}-{mesice_txt}'


@dataclass
class ImportStats:
    parsed: int = 0
    created: int = 0
    skipped_existing: int = 0
    skipped_user: int = 0
    skipped_store: int = 0
    errors: int = 0

    def as_dict(self):
        return self.__dict__.copy()


def apply_parsed_shifts(
    shifts,
    *,
    rok=2026,
    mesice=(3, 4, 5),
    dry_run=True,
    replace=False,
    replace_period=False,
):
    user_lookup, user_dupes = build_user_lookup()
    shifts = filter_brigadnik_only_users(shifts, user_lookup)
    store_lookup = build_prodejna_lookup()
    stats = ImportStats(parsed=len(shifts))
    poznamka = import_poznamka(rok, mesice)

    if replace_period:
        period_qs = Smena.objects.filter(
            datum__year=rok,
            datum__month__in=mesice,
            aktivni=True,
        )
        if not dry_run:
            period_qs.delete()
    elif replace and not dry_run:
        Smena.objects.filter(
            poznamka__startswith=IMPORT_POZNAMKA_PREFIX,
            datum__year=rok,
            datum__month__in=mesice,
        ).delete()

    to_create = []
    for shift in shifts:
        user = user_lookup.get(shift.surname_key)
        if not user:
            stats.skipped_user += 1
            continue

        prodejna_obj = None
        if shift.typ_smeny == 'prace':
            prodejna_obj = store_lookup.get(shift.prodejna_nazev.lower())
            if not prodejna_obj:
                try:
                    prodejna_obj = resolve_prodejna(shift.prodejna_nazev, shift.typ_smeny)
                except Exception:
                    stats.skipped_store += 1
                    continue

        cas_od, cas_do = shift.cas_od, shift.cas_do
        if is_absence_shift(shift.typ_smeny):
            cas_od, cas_do = normalize_dovolena_casy(shift.datum, cas_od, cas_do)
            cas_od = datetime.strptime(cas_od, '%H:%M').time()
            cas_do = datetime.strptime(cas_do, '%H:%M').time()

        brigadnik_rezim = shift.brigadnik_rezim
        if shift.typ_smeny == 'prace' and getattr(user, 'role', None) == 'BRIGADNIK':
            brigadnik_rezim = shift.brigadnik_rezim if shift.brigadnik_rezim in ('prodejce', 'vypomoc') else 'vypomoc'
        else:
            brigadnik_rezim = 'prodejce'

        pozice = shift.pozice_smeny
        if is_backoffice_user(user):
            pozice = 'backoffice'
        elif shift.typ_smeny != 'prace' or not prodejna_obj or not prodejna_obj.povolena_pozice_servis:
            pozice = 'prodej'

        note_parts = [poznamka, shift.source_file]
        if shift.poznamka_extra:
            note_parts.append(shift.poznamka_extra)
        full_poznamka = ' | '.join(note_parts)

        if not replace_period and find_overlapping_shift(
            user, shift.datum, prodejna_obj, shift.typ_smeny, cas_od, cas_do,
        ):
            stats.skipped_existing += 1
            continue

        if dry_run:
            stats.created += 1
            continue

        to_create.append(Smena(
            user=user,
            prodejna=prodejna_obj,
            datum=shift.datum,
            cas_od=cas_od,
            cas_do=cas_do,
            typ_smeny=shift.typ_smeny,
            brigadnik_rezim=brigadnik_rezim,
            pozice_smeny=pozice,
            poznamka=full_poznamka,
            aktivni=True,
        ))

    if not dry_run and to_create:
        with transaction.atomic():
            Smena.objects.bulk_create(to_create, batch_size=500)
            stats.created = len(to_create)

    return stats, user_dupes


def compare_monthly_hours_with_json(rok=2026, mesice=(3, 4, 5), tolerance=0.5):
    from shifts.payroll_service import _odpracovano_h_mesic
    from shifts.prumer_mzdy_override import load_prumer_mzdy_overrides

    overrides = load_prumer_mzdy_overrides()
    user_lookup, _ = build_user_lookup()
    rows = []
    for surname, data in sorted(overrides.items()):
        user = user_lookup.get(surname)
        if not user:
            continue
        mesice_data = data.get('mesice', []) if isinstance(data, dict) else data
        for entry in mesice_data:
            if entry.get('rok') != rok or entry.get('mesic') not in mesice:
                continue
            json_h = float(entry.get('odpracovano_h') or 0)
            smeny_h = float(_odpracovano_h_mesic(user.id, rok, entry['mesic']))
            diff = round(json_h - smeny_h, 2)
            rows.append({
                'surname': surname,
                'mesic': entry['mesic'],
                'json_h': json_h,
                'smeny_h': smeny_h,
                'diff': diff,
                'ok': abs(diff) <= tolerance,
            })
    return rows
