"""Import reklamační evidence z Mastersheet Excel nebo bootstrap JSON."""

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from reklamace.models import ReklamacePolozka, ReklamaceStatus

DEFAULT_EXCEL = Path.home() / 'Downloads' / 'Mastersheet - prodejny.xlsx'
BOOTSTRAP_JSON = Path(__file__).resolve().parents[2] / 'bootstrap' / 'mastersheet.json'

SHEETS = [
    ('Servis Reklamace', 'Servis'),
    ('Přerov Servis Reklamace', 'Přerov'),
    ('Šternberk Servis Reklamace', 'Šternberk'),
    ('Čepkov Servis Reklamace', 'Čepkov'),
]


def _cell(row, idx, default=None):
    return row[idx] if idx < len(row) else default


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        text = val.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _str_field(val):
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _column_map(header_row):
    """Servis list má jiné pořadí sloupců než per-prodejna listy."""
    labels = [_str_field(c).lower() for c in header_row]
    if 'název zboží' in labels and labels.index('název zboží') == 1:
        return {
            'nase_znacka': 0,
            'nazev_zbozi': 1,
            'dodavatel': 2,
            'faktura': 3,
            'ean': 4,
            'p_kod': 5,
            'datum_odeslani': 6,
            'cislo_zasilky': 7,
            'jejich_oznaceni': 8,
            'poznamka': 9,
        }
    return {
        'nase_znacka': 0,
        'jejich_oznaceni': 1,
        'nazev_zbozi': 2,
        'dodavatel': 3,
        'faktura': 4,
        'ean': 5,
        'p_kod': 6,
        'datum_odeslani': 7,
        'cislo_zasilky': 8,
        'poznamka': 9,
    }


def _row_to_item(row, cols, prodejna):
    znacka = _cell(row, cols['nase_znacka'])
    if not znacka or str(znacka).strip() == 'Naše značka':
        return None
    item = {
        'nase_znacka': str(znacka).strip(),
        'jejich_oznaceni': _str_field(_cell(row, cols['jejich_oznaceni'])),
        'nazev_zbozi': _str_field(_cell(row, cols['nazev_zbozi'])),
        'dodavatel': _str_field(_cell(row, cols['dodavatel'])),
        'faktura': _str_field(_cell(row, cols['faktura'])),
        'ean': _str_field(_cell(row, cols['ean'])),
        'p_kod': _str_field(_cell(row, cols['p_kod'])),
        'datum_odeslani': _parse_date(_cell(row, cols['datum_odeslani'])),
        'cislo_zasilky': _str_field(_cell(row, cols['cislo_zasilky'])),
        'poznamka': _str_field(_cell(row, cols['poznamka'])),
        'prodejna': prodejna,
    }
    if not item['nazev_zbozi'] and not item['dodavatel'] and not item['datum_odeslani']:
        return None
    return item


def parse_excel(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items = []
    for sheet_name, prodejna in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cols = None
        for row in ws.iter_rows(values_only=True):
            if cols is None:
                if row and _str_field(row[0]) == 'Naše značka':
                    cols = _column_map(row)
                continue
            item = _row_to_item(row, cols, prodejna)
            if item:
                items.append(item)
    wb.close()
    return items


class Command(BaseCommand):
    help = 'Import reklamační evidence z Mastersheet.'

    def add_arguments(self, parser):
        parser.add_argument('--excel', type=str, default='')
        parser.add_argument('--json', type=str, default='')
        parser.add_argument('--dry-run', action='store_true')
        parser.add_argument('--clear', action='store_true')

    def handle(self, *args, **options):
        excel_path = Path(options['excel']) if options['excel'] else DEFAULT_EXCEL
        json_path = Path(options['json']) if options['json'] else BOOTSTRAP_JSON

        if excel_path.is_file():
            items = parse_excel(excel_path)
            self.stdout.write(f'Čtu Excel: {excel_path}')
        elif json_path.is_file():
            items = json.loads(json_path.read_text(encoding='utf-8'))
            for item in items:
                item['datum_odeslani'] = _parse_date(item.get('datum_odeslani'))
            items = [i for i in items if i.get('nazev_zbozi') or i.get('dodavatel') or i.get('datum_odeslani')]
            self.stdout.write(f'Čtu JSON: {json_path}')
        else:
            self.stderr.write(self.style.ERROR('Nenalezen Excel ani bootstrap JSON.'))
            return

        self.stdout.write(f'Nalezeno {len(items)} položek')

        if options['dry_run']:
            for item in items[:3]:
                self.stdout.write(f'  {item["nase_znacka"]} {item["nazev_zbozi"][:40]}')
            self.stdout.write(self.style.WARNING('DRY RUN'))
            return

        with transaction.atomic():
            if options['clear']:
                deleted, _ = ReklamacePolozka.objects.all().delete()
                self.stdout.write(f'Smazáno {deleted}')

            created = 0
            for item in items:
                defaults = dict(item)
                if defaults.get('datum_odeslani'):
                    defaults['status'] = ReklamaceStatus.ODESLANE
                _, was_created = ReklamacePolozka.objects.get_or_create(
                    nase_znacka=item['nase_znacka'],
                    defaults=defaults,
                )
                if was_created:
                    created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Hotovo: {created} nových, celkem {ReklamacePolozka.objects.count()}'
        ))
