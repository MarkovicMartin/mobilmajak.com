"""Porovnání Excel vs systém – průměr dovolené (hodiny z Excelu vs hodiny ze směn)."""
from pathlib import Path

from django.core.management.base import BaseCommand

from shifts.payroll_service import dovolena_body_vypocet, prumer_fixni_hodinove_detail
from shifts.prumer_mzdy_override import prumer_override_for_user
from shifts.vacation_service import dovolena_stav, is_dovolena_eligible
from users.models import WebUser

VACATION_KEYS = [
    'králik', 'kováčik', 'valenta', 'gabriel', 'babušík', 'kolarčík',
    'vychodil', 'markovič', 'krumpolc', 'létal', 'karas', 'hekele',
]


def _load_excel_list14(path):
    import openpyxl

    def num(x):
        if x is None or (isinstance(x, str) and x.startswith('#')):
            return 0.0
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['List 14']
    out = {}
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        key = str(name).strip().lower().rstrip('?')
        out[key] = []
        for i, m in enumerate([3, 4, 5]):
            out[key].append({
                'mesic': m,
                'h': num(ws.cell(r, 2 + i * 2).value),
                'celkem': num(ws.cell(r, 3 + i * 2).value),
            })
    return out


class Command(BaseCommand):
    help = 'Porovná Excel (List 14) vs systém – hodiny, výplata, průměr dovolené.'

    def add_arguments(self, parser):
        parser.add_argument('--rok', type=int, default=2026)
        parser.add_argument('--mesic', type=int, default=6, help='Referenční měsíc průměru')
        parser.add_argument(
            '--excel',
            default='/Users/m/Downloads/Odměňování.xlsx',
            help='Cesta k Odměňování.xlsx',
        )

    def handle(self, *args, **options):
        rok = options['rok']
        ref = options['mesic']
        mesice = [ref - 3, ref - 2, ref - 1]
        excel_path = Path(options['excel'])
        if not excel_path.exists():
            self.stderr.write(f'Chybí Excel: {excel_path}')
            return

        excel = _load_excel_list14(excel_path)
        user_qs = WebUser.objects.filter(aktivni=True).only(
            'id',
            'prijmeni',
            'jmeno',
            'role',
            'technik_id',
            'mzda_zaklad',
            'mzda_doplnky',
            'mzda_cestovne',
            'dovolena_fond_extra_h',
            'dovolena_korekce_cerpano_h',
        )
        by_name = {u.prijmeni.strip().lower(): u for u in user_qs}
        for u in WebUser.objects.filter(aktivni=True, prijmeni__iexact='Markovič').only('id', 'prijmeni', 'jmeno', 'role', 'technik_id'):
            by_name['markovič'] = u

        self.stdout.write(f'Březen–Květen {rok} | průměr dovolené pro {ref:02d}/{rok}')
        self.stdout.write('Σ Excel = CELKEM NOVE | Σ fix = fixní část z profilu')
        self.stdout.write('')
        self.stdout.write(
            f\"{'Příjmení':<11} {'H X':>5} {'H sys':>6} {'Σ Excel':>8} {'Σ fix':>8} \"
            f\"{'sazbaX':>6} {'sazbaS':>6} {'Δ':>5} {'dov.h':>5} {'bodX':>6} {'bodS':>6}\"
        )
        self.stdout.write('-' * 88)

        for key in VACATION_KEYS:
            user = by_name.get(key)
            if not user:
                self.stdout.write(f'{key:<11} — uživatel nenalezen')
                continue
            if not is_dovolena_eligible(user):
                self.stdout.write(f'{user.prijmeni:<11} — bez nároku na dovolenou')
                continue

            ex = excel.get(key, [])
            override = prumer_override_for_user(user)
            dx = prumer_fixni_hodinove_detail(user, rok, ref, override_mesice=override)
            ds = prumer_fixni_hodinove_detail(user, rok, ref, override_mesice=None)
            sum_excel = sum(m['celkem'] for m in ex)
            sum_fix = dx['celkem_fixni']
            stav = dovolena_stav(user, rok) or {}
            dovh = float(stav.get('cerpano_smeny_h') or 0)
            sx, ss = dx['prumer_fixni_h'], ds['prumer_fixni_h']
            bx = float(dovolena_body_vypocet(user, dovh, sx))
            bs = float(dovolena_body_vypocet(user, dovh, ss))
            self.stdout.write(
                f"{user.prijmeni:<11} {dx['celkem_h']:>5.0f} {ds['celkem_h']:>6.0f} "
                f"{sum_excel:>8.0f} {sum_fix:>8.0f} "
                f"{sx:>6.0f} {ss:>6.0f} {sx - ss:>+5.0f} "
                f"{dovh:>5.0f} {bx:>6.0f} {bs:>6.0f}"
            )

        self.stdout.write('')
        self.stdout.write('Měsíční detail (Excel vs směny):')
        for key in VACATION_KEYS:
            user = by_name.get(key)
            if not user or not is_dovolena_eligible(user):
                continue
            exm = {m['mesic']: m for m in excel.get(key, [])}
            override = prumer_override_for_user(user)
            dx = prumer_fixni_hodinove_detail(user, rok, ref, override_mesice=override)
            parts = []
            for m in dx['mesice']:
                em = exm.get(m['mesic'], {})
                h_sm = m.get('odpracovano_h_smeny', 0)
                if abs(m.get('hodiny_rozdil_h', 0)) > 0.5 or em.get('celkem'):
                    parts.append(
                        f"{m['mesic']:02d}: Excel {em.get('celkem', 0):.0f}b/{m['odpracovano_h']:.0f}h "
                        f"směny {h_sm:.0f}h"
                    )
            if parts:
                self.stdout.write(f"  {user.prijmeni}: " + ' | '.join(parts))
